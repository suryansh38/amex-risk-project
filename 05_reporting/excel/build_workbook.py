"""
Builds an analyst-facing Excel workbook from the outputs of the credit,
fraud, and portfolio risk modules -- with live formulas (not just pasted
values), matching what a risk analyst actually hands to a manager: a
workbook they can re-sort, re-pivot, and audit the formulas of.

Run (after the upstream modules have produced their output/ CSVs -- falls
back to synthetic placeholders for any module not yet run):
    python build_workbook.py
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

ROOT = Path(__file__).resolve().parents[2]
OUT_XLSX = Path(__file__).resolve().parent / "amex_risk_summary.xlsx"

HEADER_FILL = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    for j, col in enumerate(df.columns, start=1):
        width = max(12, min(28, int(df[col].astype(str).str.len().max() if len(df) else 10) + 2))
        ws.column_dimensions[get_column_letter(j)].width = width
    return start_row + len(df) + 1


def load_or_placeholder(path: Path, placeholder: pd.DataFrame) -> pd.DataFrame:
    return pd.read_csv(path) if path.exists() else placeholder


def main() -> None:
    credit_out = ROOT / "02_credit_risk" / "python" / "output"
    portfolio_out = ROOT / "04_portfolio_risk" / "output"
    fraud_out = ROOT / "03_fraud_risk" / "output"

    book_summary = load_or_placeholder(
        portfolio_out / "book_summary.csv",
        pd.DataFrame([{"n_accounts": 0, "total_exposure": 0, "total_expected_loss": 0,
                        "el_rate_pct": 0, "avg_pd_pct": 0}]),
    )
    decile = load_or_placeholder(
        portfolio_out / "expected_loss_by_decile.csv",
        pd.DataFrame(columns=["pd_decile", "n_accounts", "exposure", "expected_loss", "avg_pd_pct"]),
    )
    stress = load_or_placeholder(
        portfolio_out / "stress_test_results.csv",
        pd.DataFrame(columns=["scenario", "pd_multiplier", "stressed_el", "el_rate_pct"]),
    )
    model_metrics = load_or_placeholder(
        credit_out / "model_metrics.csv",
        pd.DataFrame([{"auc": None, "ks": None, "gini": None, "n_test": None}]),
    )
    fraud_metrics = load_or_placeholder(
        fraud_out / "model_metrics.csv",
        pd.DataFrame([{"roc_auc": None, "pr_auc": None, "operating_threshold": None, "n_test": None}]),
    )

    wb = Workbook()

    # --- Sheet 1: Executive Summary -----------------------------------
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = "Amex Risk Analytics — Portfolio Summary"
    ws["A1"].font = Font(size=14, bold=True, color="1A365D")
    write_df(ws, book_summary, start_row=3)

    ws["A8"] = "Credit Model Performance (Python/XGBoost)"
    ws["A8"].font = Font(bold=True)
    write_df(ws, model_metrics, start_row=9)

    ws["A13"] = "Fraud Model Performance (Python/RandomForest)"
    ws["A13"].font = Font(bold=True)
    write_df(ws, fraud_metrics, start_row=14)

    # --- Sheet 2: Expected Loss by Decile (with live formula + chart) ---
    ws2 = wb.create_sheet("Expected Loss by Decile")
    end_row = write_df(ws2, decile, start_row=1)
    if len(decile) > 0:
        n = len(decile)
        # live-formula total row instead of a pasted value
        ws2.cell(row=end_row + 1, column=1, value="Total").font = Font(bold=True)
        for col_idx, col_name in enumerate(decile.columns, start=1):
            if col_name in ("n_accounts", "exposure", "expected_loss"):
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}2:{col_letter}{n + 1})"
                ws2.cell(row=end_row + 1, column=col_idx, value=formula).font = Font(bold=True)

        chart = BarChart()
        chart.title = "Expected Loss by PD Decile"
        chart.y_axis.title = "Expected Loss ($)"
        chart.x_axis.title = "PD Decile"
        data = Reference(ws2, min_col=decile.columns.get_loc("expected_loss") + 1,
                          min_row=1, max_row=n + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=n + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws2.add_chart(chart, f"A{end_row + 4}")

    # --- Sheet 3: Stress Test Results -----------------------------------
    ws3 = wb.create_sheet("Stress Test")
    write_df(ws3, stress, start_row=1)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
